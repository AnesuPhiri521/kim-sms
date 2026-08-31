"""Render the Financial Reports (doc 08) to CSV / XLSX / PDF downloads.

Every report on the reports screen is ultimately a titled table of rows, so
one renderer handles all of them. Callers build a `columns` spec + `rows`
(money kept as integer cents, exactly like the rest of the fee module) and
pick a format; this module owns the per-format formatting so "USD 1,234.50"
in the PDF and a real `1234.50` number in the spreadsheet stay in sync.
"""

import csv
import io
from dataclasses import dataclass
from datetime import date
from enum import StrEnum
from typing import Any, NamedTuple

from app.db.base import utcnow


class ExportFormat(StrEnum):
    csv = "csv"
    xlsx = "xlsx"
    pdf = "pdf"


class ExportedFile(NamedTuple):
    content: bytes
    media_type: str
    filename: str


# Column value kinds — drive per-format rendering.
TEXT = "text"
MONEY = "money"  # value is an integer number of cents
NUMBER = "number"
PERCENT = "percent"  # value is a number like 87.5


@dataclass(frozen=True)
class Column:
    header: str
    kind: str = TEXT


_MEDIA_TYPES = {
    ExportFormat.csv: "text/csv; charset=utf-8",
    ExportFormat.xlsx: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ExportFormat.pdf: "application/pdf",
}


def _dollars(cents: Any) -> float:
    return round((cents or 0) / 100, 2)


def render_report(
    fmt: ExportFormat,
    *,
    slug: str,
    title: str,
    meta_lines: list[str],
    columns: list[Column],
    rows: list[list[Any]],
    currency_code: str,
    total_row: list[Any] | None = None,
) -> ExportedFile:
    generated = f"Generated {utcnow().strftime('%d %b %Y %H:%M UTC')}"
    filename = f"{slug}-{date.today().isoformat()}.{fmt.value}"
    all_meta = [*meta_lines, generated]
    if fmt is ExportFormat.csv:
        content = _render_csv(columns, rows, total_row)
    elif fmt is ExportFormat.xlsx:
        content = _render_xlsx(title, all_meta, columns, rows, total_row)
    else:
        content = _render_pdf(title, all_meta, columns, rows, currency_code, total_row)
    return ExportedFile(content=content, media_type=_MEDIA_TYPES[fmt], filename=filename)


# --------------------------------------------------------------------- csv --


def _csv_cell(col: Column, value: Any) -> Any:
    if value is None:
        return ""
    if col.kind == MONEY:
        return f"{_dollars(value):.2f}"
    if col.kind == PERCENT:
        return f"{value:.2f}"
    return value


def _render_csv(columns: list[Column], rows: list[list[Any]], total_row: list[Any] | None) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([c.header for c in columns])
    for row in rows:
        writer.writerow([_csv_cell(c, v) for c, v in zip(columns, row, strict=False)])
    if total_row is not None:
        writer.writerow([_csv_cell(c, v) for c, v in zip(columns, total_row, strict=False)])
    # BOM so Excel opens the file as UTF-8 rather than the system code page.
    return buffer.getvalue().encode("utf-8-sig")


# -------------------------------------------------------------------- xlsx --


def _render_xlsx(
    title: str,
    meta_lines: list[str],
    columns: list[Column],
    rows: list[list[Any]],
    total_row: list[Any] | None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    sheet.append([title])
    sheet["A1"].font = Font(bold=True, size=14)
    for line in meta_lines:
        sheet.append([line])
    sheet.append([])

    header_idx = sheet.max_row + 1
    sheet.append([c.header for c in columns])
    for cell in sheet[header_idx]:
        cell.font = Font(bold=True)

    def append_row(values: list[Any], *, bold: bool = False) -> None:
        out: list[Any] = []
        for col, value in zip(columns, values, strict=False):
            if col.kind == MONEY:
                out.append(_dollars(value))
            elif col.kind in (NUMBER, PERCENT):
                out.append(value if value is not None else "")
            else:
                out.append("" if value is None else value)
        sheet.append(out)
        r = sheet.max_row
        for i, col in enumerate(columns, start=1):
            if col.kind == MONEY:
                sheet.cell(r, i).number_format = "#,##0.00"
            elif col.kind == PERCENT:
                sheet.cell(r, i).number_format = '0.00"%"'
            if bold:
                sheet.cell(r, i).font = Font(bold=True)

    for row in rows:
        append_row(row)
    if total_row is not None:
        append_row(total_row, bold=True)

    for i, col in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = max(16, len(col.header) + 4)

    out = io.BytesIO()
    workbook.save(out)
    return out.getvalue()


# --------------------------------------------------------------------- pdf --


def _pdf_cell(col: Column, value: Any, currency_code: str) -> str:
    if value is None or value == "":
        return ""
    if col.kind == MONEY:
        return f"{currency_code} {_dollars(value):,.2f}"
    if col.kind == PERCENT:
        return f"{value:.2f}%"
    return str(value)


def _render_pdf(
    title: str,
    meta_lines: list[str],
    columns: list[Column],
    rows: list[list[Any]],
    currency_code: str,
    total_row: list[Any] | None,
) -> bytes:
    from fpdf import FPDF

    pdf = FPDF(orientation="L" if len(columns) > 4 else "P")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_margins(14, 14, 14)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 15)
    pdf.cell(0, 9, title, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    for line in meta_lines:
        pdf.cell(0, 5, line, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    if not rows:
        pdf.set_font("Helvetica", "I", 10)
        pdf.cell(0, 6, "No data for the selected filters.", new_x="LMARGIN", new_y="NEXT")
        return bytes(pdf.output())

    align = tuple("RIGHT" if c.kind in (MONEY, NUMBER, PERCENT) else "LEFT" for c in columns)
    pdf.set_font("Helvetica", "", 10)
    with pdf.table(text_align=align, padding=1.5) as table:
        header = table.row()
        for col in columns:
            header.cell(col.header)
        for row in rows:
            line = table.row()
            for col, value in zip(columns, row, strict=False):
                line.cell(_pdf_cell(col, value, currency_code))
        if total_row is not None:
            line = table.row()
            for col, value in zip(columns, total_row, strict=False):
                line.cell(_pdf_cell(col, value, currency_code))

    return bytes(pdf.output())
